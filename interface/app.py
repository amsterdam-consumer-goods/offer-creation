"""
MAIN APPLICATION INTERFACE
--------------------------
Streamlit UI for the Offer Creation Tool. Handles:
- File upload and validation (Excel, Images, PDF)
- Department selection (FOOD/HPC)
- Processing orchestration
- Results display and download
"""

import tempfile
from pathlib import Path

import numpy as np
import pandas as pd
import streamlit as st

from components import (
    render_department_selector,
    render_download_buttons,
    render_file_uploader,
    render_header,
    render_process_button,
    render_product_image_uploader,
    render_reset_button,
    render_selectable_table,
    render_success_message,
)
from processor import process_uploaded_file
from styles import get_custom_css

import sys
from pathlib import Path
sys.path.append(str(Path(__file__).parent.parent))

from config import (
    MAX_FILE_SIZE_MB,
    MAX_SHEET_ROWS,
    MAX_SHEET_COLS,
    MAX_SHEETS,
    EXTREME_COLS_LIMIT,
)


def _force_availability_ints(df: pd.DataFrame) -> pd.DataFrame:
    """Ensure availability columns display as integers (no .0 decimals)."""
    if df is None or df.empty:
        return df

    cols = ["Availability/Cartons", "Availability/Pieces", "Availability/Pallets"]
    for c in cols:
        if c in df.columns:
            s = pd.to_numeric(df[c], errors="coerce")
            s = np.ceil(s)
            df[c] = s.astype("Int64")
    return df


def _get_file_type(uploaded_file) -> str:
    """
    Detect file type from extension.
    
    Returns: 'excel', 'image', 'pdf', or 'unknown'
    """
    filename = uploaded_file.name.lower()
    
    if filename.endswith(('.xlsx', '.xls')):
        return 'excel'
    elif filename.endswith(('.jpg', '.jpeg', '.png', '.gif', '.webp')):
        return 'image'
    elif filename.endswith('.pdf'):
        return 'pdf'
    else:
        return 'unknown'


def _validate_excel_file(uploaded_file) -> tuple[bool, str, dict]:
    """
    Validate Excel file size and structure.
    
    Returns:
        (is_valid, error_message, sheet_info_dict)
        sheet_info_dict = {sheet_name: {"rows": int, "cols": int}}
    """
    import openpyxl
    
    file_size_mb = uploaded_file.size / (1024 * 1024)
    if file_size_mb > MAX_FILE_SIZE_MB:
        return False, f"❌ File size ({file_size_mb:.1f} MB) exceeds limit ({MAX_FILE_SIZE_MB} MB). Please reduce file size.", {}
    
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            tmp.write(uploaded_file.getbuffer())
            tmp_path = tmp.name
        
        wb = openpyxl.load_workbook(tmp_path, read_only=True, data_only=True)
        
        sheet_info = {}
        total_sheets = len(wb.sheetnames)
        
        if total_sheets > MAX_SHEETS:
            wb.close()
            Path(tmp_path).unlink(missing_ok=True)
            return False, f"❌ File has {total_sheets} sheets. Maximum {MAX_SHEETS} sheets allowed.", {}
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = ws.max_row
            cols = ws.max_column
            
            sheet_info[sheet_name] = {"rows": rows, "cols": cols}
        
        wb.close()
        Path(tmp_path).unlink(missing_ok=True)
        
        return True, "", sheet_info
        
    except Exception as e:
        return False, f"❌ Error reading Excel file: {str(e)}", {}


def _check_sheet_limits(sheet_info: dict, selected_sheet: str) -> tuple[bool, str]:
    """
    Check if selected sheet exceeds processing limits.
    
    Returns:
        (is_valid, error_message)
    """
    if selected_sheet not in sheet_info:
        return False, "❌ Selected sheet not found in file."
    
    info = sheet_info[selected_sheet]
    rows = info["rows"]
    cols = info["cols"]
    
    if cols > EXTREME_COLS_LIMIT:
        return False, f"❌ Sheet has {cols:,} columns (limit: {EXTREME_COLS_LIMIT:,}). This is too wide to process. Please split the data."
    
    if rows > MAX_SHEET_ROWS:
        return False, f"❌ Sheet has {rows:,} rows (limit: {MAX_SHEET_ROWS:,}). Please filter or split the data."
    
    if cols > MAX_SHEET_COLS:
        return False, f"❌ Sheet has {cols:,} columns (limit: {MAX_SHEET_COLS:,}). Please reduce columns."
    
    return True, ""


# ============================================================================
# PAGE CONFIG
# ============================================================================
st.set_page_config(
    page_title="Offer Creator",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="collapsed",
)

st.markdown(get_custom_css(), unsafe_allow_html=True)

# ============================================================================
# SESSION STATE INITIALIZATION
# ============================================================================
if "processed" not in st.session_state:
    st.session_state.processed = False
if "output_path" not in st.session_state:
    st.session_state.output_path = None
if "df" not in st.session_state:
    st.session_state.df = None
if "selected_df" not in st.session_state:
    st.session_state.selected_df = None
if "row_selected" not in st.session_state:
    st.session_state.row_selected = None
if "double_stackable" not in st.session_state:
    st.session_state.double_stackable = False
if "extract_price" not in st.session_state:
    st.session_state.extract_price = False
if "product_images" not in st.session_state:
    st.session_state.product_images = {}
if "uploaded_file_data" not in st.session_state:
    st.session_state.uploaded_file_data = None
if "dept_type" not in st.session_state:
    st.session_state.dept_type = None
if "sheet_info" not in st.session_state:
    st.session_state.sheet_info = {}
if "selected_sheet" not in st.session_state:
    st.session_state.selected_sheet = None
if "file_validated" not in st.session_state:
    st.session_state.file_validated = False
if "file_type" not in st.session_state:
    st.session_state.file_type = None

# ============================================================================
# MAIN APP FLOW
# ============================================================================
render_header()

dept_type, double_stackable, extract_price = render_department_selector()
st.session_state.double_stackable = double_stackable
st.session_state.extract_price = extract_price
st.session_state.dept_type = dept_type

if dept_type:
    uploaded_file = render_file_uploader()

    if uploaded_file:
        st.session_state.uploaded_file_data = uploaded_file
        
        file_is_new = st.session_state.get("last_file_name") != uploaded_file.name
        
        if file_is_new or not st.session_state.file_validated:
            file_type = _get_file_type(uploaded_file)
            st.session_state.file_type = file_type
            
            # Reject unsupported formats
            if file_type == 'unknown':
                st.error("❌ **Unsupported file format**")
                st.info("📄 **Supported formats:**\n"
                       "- Excel: .xlsx, .xls\n"
                       "- Images: .jpg, .jpeg, .png\n"
                       "- PDF: .pdf")
                st.stop()
            
            # Excel: Full validation with sheet limits
            if file_type == 'excel':
                with st.spinner("🔍 Checking Excel file limits..."):
                    is_valid, error_msg, sheet_info = _validate_excel_file(uploaded_file)
                    
                    if not is_valid:
                        st.error(error_msg)
                        st.info("💡 **How to reduce file size:**\n"
                               "- Filter data to only include relevant rows\n"
                               "- Remove unnecessary columns\n"
                               "- Split large files into smaller batches")
                        st.stop()
                    
                    first_sheet = list(sheet_info.keys())[0]
                    
                    is_valid, limit_error = _check_sheet_limits(sheet_info, first_sheet)
                    if not is_valid:
                        info = sheet_info[first_sheet]
                        st.error(f"❌ **File is too large to process**\n\n"
                                f"Sheet '{first_sheet}' has **{info['rows']:,} rows** and **{info['cols']:,} columns**.\n\n"
                                f"**Limits:** {MAX_SHEET_ROWS:,} rows, {MAX_SHEET_COLS:,} columns per sheet.")
                        st.info("💡 **How to fix:**\n"
                               f"- Filter the sheet to reduce rows (currently {info['rows']:,} → target <{MAX_SHEET_ROWS:,})\n"
                               f"- Remove unused columns (currently {info['cols']:,} → target <{MAX_SHEET_COLS:,})\n"
                               "- Split the file into multiple smaller files")
                        st.stop()
                    
                    st.session_state.sheet_info = sheet_info
                    st.session_state.selected_sheet = first_sheet
                    st.session_state.file_validated = True
                    st.session_state.last_file_name = uploaded_file.name
                    
                    info = sheet_info[first_sheet]
                    st.success(f"✅ Excel file validated! Sheet '{first_sheet}' ({info['rows']:,} rows × {info['cols']:,} cols)")
            
            # Image: Simple size check
            elif file_type == 'image':
                file_size_mb = uploaded_file.size / (1024 * 1024)
                if file_size_mb > MAX_FILE_SIZE_MB:
                    st.error(f"❌ Image size ({file_size_mb:.1f} MB) exceeds limit ({MAX_FILE_SIZE_MB} MB)")
                    st.stop()
                
                st.session_state.file_validated = True
                st.session_state.last_file_name = uploaded_file.name
                st.success(f"✅ Image uploaded! ({file_size_mb:.2f} MB) - AI will extract data from image")
            
            # PDF: Simple size check
            elif file_type == 'pdf':
                file_size_mb = uploaded_file.size / (1024 * 1024)
                if file_size_mb > MAX_FILE_SIZE_MB:
                    st.error(f"❌ PDF size ({file_size_mb:.1f} MB) exceeds limit ({MAX_FILE_SIZE_MB} MB)")
                    st.stop()
                
                st.session_state.file_validated = True
                st.session_state.last_file_name = uploaded_file.name
                st.success(f"✅ PDF uploaded! ({file_size_mb:.2f} MB) - AI will extract data from PDF")

        # ====================================================================
        # PROCESS BUTTON
        # ====================================================================
        
        if st.session_state.file_validated:
            process_btn = render_process_button()

            if process_btn:
                with st.spinner("🔄 Processing your offer..."):
                    success, output_path, df, error = process_uploaded_file(
                        uploaded_file=uploaded_file,
                        dept_type=dept_type,
                        double_stackable=st.session_state.double_stackable,
                        extract_price=st.session_state.extract_price,
                        product_images=None,
                        selected_sheet=st.session_state.selected_sheet,
                    )

                    if success:
                        df = _force_availability_ints(df)

                        st.session_state.processed = True
                        st.session_state.output_path = output_path
                        st.session_state.df = df
                        st.session_state.selected_df = None
                        st.session_state.product_images = {}
                        st.session_state.row_selected = None
                    else:
                        st.error(f"❌ Error: {error}")

# ============================================================================
# RESULTS SECTION
# ============================================================================
if st.session_state.processed and st.session_state.df is not None:
    render_success_message()

    edited_df = render_selectable_table(st.session_state.df)

    if edited_df is not None and len(edited_df) > 0 and "Include" in edited_df.columns:
        selected_df = edited_df[edited_df["Include"] == True].copy()
        selected_df.drop(columns=["Include"], inplace=True, errors="ignore")
        selected_df.reset_index(drop=True, inplace=True)
    else:
        selected_df = edited_df.copy() if edited_df is not None else None
        if selected_df is not None:
            selected_df.reset_index(drop=True, inplace=True)

    if selected_df is not None:
        selected_df = _force_availability_ints(selected_df)

    st.session_state.selected_df = selected_df

    product_images = render_product_image_uploader(selected_df)
    st.session_state.product_images = product_images

    action, images_to_use = render_download_buttons(
        selected_df,
        product_images,
        dept_type=st.session_state.dept_type,
        base_filename=st.session_state.output_path.name,
    )

    if action == "no_images" and selected_df is not None and len(selected_df) > 0:
        with st.spinner("📄 Generating Excel (no images)..."):
            from domain.schemas import FOOD_HEADERS, HPC_HEADERS
            from writers.excel_writer import write_rows_to_xlsx

            headers = FOOD_HEADERS if st.session_state.dept_type == "food" else HPC_HEADERS
            rows = selected_df.to_dict(orient="records")

            base = st.session_state.output_path.name if st.session_state.output_path else "offer.xlsx"
            output_no_images = Path(tempfile.gettempdir()) / f"no_images_{base}"

            write_rows_to_xlsx(
                output_path=output_no_images,
                sheet_name=st.session_state.dept_type.upper(),
                headers=headers,
                rows=rows,
                product_images=None,
            )

            st.success("✅ Excel (no images) generated!")

            with open(output_no_images, "rb") as f:
                st.download_button(
                    label="📥 Download Excel (No Images)",
                    data=f,
                    file_name=output_no_images.name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="secondary",
                    width="stretch",
                    key="download_no_images",
                )

    if action == "with_images" and images_to_use and selected_df is not None and len(selected_df) > 0:
        with st.spinner("🎨 Generating Excel with images..."):
            from domain.schemas import FOOD_HEADERS, HPC_HEADERS
            from writers.excel_writer import write_rows_to_xlsx

            headers = FOOD_HEADERS if st.session_state.dept_type == "food" else HPC_HEADERS

            temp_dir = Path(tempfile.gettempdir()) / "offer_images"
            temp_dir.mkdir(exist_ok=True)

            image_paths: list[Path | None] = []
            for idx in range(len(selected_df)):
                if idx in images_to_use:
                    img_file = images_to_use[idx]
                    img_ext = Path(img_file.name).suffix
                    img_path = temp_dir / f"product_{idx}{img_ext}"

                    with open(img_path, "wb") as f:
                        f.write(img_file.getbuffer())

                    image_paths.append(img_path)
                else:
                    image_paths.append(None)

            rows = selected_df.to_dict(orient="records")

            base = st.session_state.output_path.name if st.session_state.output_path else "offer.xlsx"
            output_with_images = Path(tempfile.gettempdir()) / f"with_images_{base}"

            write_rows_to_xlsx(
                output_path=output_with_images,
                sheet_name=st.session_state.dept_type.upper(),
                headers=headers,
                rows=rows,
                product_images=image_paths,
            )

            st.success("✅ Excel with images generated!")

            with open(output_with_images, "rb") as f:
                st.download_button(
                    label="📥 Download Excel with Images",
                    data=f,
                    file_name=output_with_images.name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary",
                    width="stretch",
                    key="final_download",
                )

    if render_reset_button():
        for key in list(st.session_state.keys()):
            del st.session_state[key]
        st.rerun()